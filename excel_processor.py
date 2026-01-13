# ============================
# excel_processor.py
# ============================

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from shutil import copy2
from helpers.paths import DATA_FOLDER


# -----------------------------
# Constants
# -----------------------------
EXCEL_RESULTS_FILENAME = "results" # Name of the .xlsx file to store results (e.g., "results")
EXCEL_RESULTS_EXTENSION = ".xlsx" # Excel file extension
EXCEL_BACKUP_FOLDER_NAME = "excel_backup" # Folder name for backups

# -----------------------------
# Excel helper functions
# -----------------------------

def get_excel_path(): 
    """Returns the Excel excel_path."""
    return os.path.join(DATA_FOLDER, f"{EXCEL_RESULTS_FILENAME}{EXCEL_RESULTS_EXTENSION}") # Full path to results.xlsx (e.g., .../data/results.xlsx)

def get_current_patient(viewer):
    return viewer.current_patient

def get_button_status(viewer):
    """
    Reads the 5 toggle buttons (0/1) and dominance (0/1/2)
    returns a list of 6 values.
    """

    toggle_states = [
        viewer.ids.toggle_1.state,
        viewer.ids.toggle_2.state,
        viewer.ids.toggle_3.state,
        viewer.ids.toggle_4.state,
        viewer.ids.toggle_5.state,
    ]

    toggle_values = [1 if t == "down" else 0 for t in toggle_states]

    dominance_value = int(viewer.dominance)  # 0,1,2

    status = toggle_values + [dominance_value]

    print("\n--- Button Status ---")
    for i, val in enumerate(status, start=1):
        print(f"Value {i}: {val}")
    print("---------------------\n")

    return status



def excel_create(excel_path):
    """Creates the Excel file with headers if it does not exist."""
    if os.path.exists(excel_path):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "id" # Patient ID
    ws["B1"] = "bad_img_qual" # Button 1 = bad image quality
    ws["C1"] = "small_seg_err" # Button 2 = small segmentation error
    ws["D1"] = "large_seg_err" # Button 3 = large segmentation error
    ws["E1"] = "artifact" # Button 4 = artifact
    ws["F1"] = "infarct" # Button 5 = infarct
    ws["G1"] = "dominance" # Dominance: 0=Right, 1=Left, 2=Codominance

    wb.save(excel_path)
    wb.close()
    print(f"Created Excel file: {excel_path}")


def excel_write(excel_path, patient_id, button_status):
    """
    Writes a row for the given patient_id.
    Overwrites that row each time.
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    row = patient_id + 1

    ws.cell(row=row, column=1, value=patient_id)

    for i, value in enumerate(button_status, start=2):
        ws.cell(row=row, column=i, value=value)

    wb.save(excel_path)
    wb.close()
    print(f"Wrote data for patient {patient_id} → Row {row}")

def excel_row_has_data(excel_path, patient_id):
    """
    Returns True if the patient row already contains data.
    Row = patient_id + 1
    """
    if not os.path.exists(excel_path):
        return False

    wb = load_workbook(excel_path)
    ws = wb.active

    row = patient_id + 1

    # Check columns B-G (toggle values)
    has_data = any(ws.cell(row=row, column=col).value is not None for col in range(2, 8))

    wb.close()
    return has_data


def excel_backup(excel_path):
    """
    Creates a backup copy of the Excel file.
    Backup excel_path format:
        e.g. excel_backup/results_20250112_153022.xlsx
    """

    # Create folder if missing
    backup_dir = os.path.join(DATA_FOLDER, EXCEL_BACKUP_FOLDER_NAME) # backup folder (e.g., .../data/excel_backup)
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir, exist_ok=True) 
        print(f"Created backup folder: {EXCEL_BACKUP_FOLDER_NAME} at {backup_dir}")

    # Format timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S") # current date and time, e.g. 20250112_153022

    # Build backup excel_path (e.g. .../data/excel_backup/results_20250112_153022.xlsx)
    backup_path = os.path.join(backup_dir, f"{EXCEL_RESULTS_FILENAME}_{timestamp}{EXCEL_RESULTS_EXTENSION}") 
    # Copy file
    if os.path.exists(excel_path):
        copy2(excel_path, backup_path)
        print(f"Backup created → {backup_path}")
    else:
        print("Backup skipped: original file does not exist yet.")

def excel_read(excel_path, patient_id):
    """
    Reads toggle values (B–G) for a given patient.
    Returns a list of 0/1 or None if row is empty.
    """
    if not os.path.exists(excel_path):
        return None

    wb = load_workbook(excel_path)
    ws = wb.active
    row = patient_id + 1

    values = []
    for col in range(2, 8): # Columns B to G
        values.append(ws.cell(row=row, column=col).value)

    wb.close()

    if all(v is None for v in values):
        return None

    return values

# -----------------------------
# Main Excel orchestrator
# -----------------------------

def excel_process(viewer):
    """
    Full pipeline:
     1) Define file name
     2) Read patient ID
     3) Read toggle values
     4) Create Excel if needed
     5) Backup if row already contains data
     6) Write row
    """
    excel_path = get_excel_path()
    patient_id = get_current_patient(viewer)
    status = get_button_status(viewer)

    # Step 1: Create the file if missing
    excel_create(excel_path)

    # Step 2: Backup BEFORE overwriting existing data
    if excel_row_has_data(excel_path, patient_id):
        print("Row already has data → creating backup BEFORE overwrite")
        excel_backup(excel_path)
    else:
        print("Row is empty → NO backup needed")

    # Step 3: Write processed values
    excel_write(excel_path, patient_id, status)
